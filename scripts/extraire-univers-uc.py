#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrait l univers des unites de compte des deux partenaires vers
public/data/univers-uc-<partenaire>.json, que l ecran « Allocations types »
charge a la demande.

Pourquoi un script et pas une saisie a la main : les listes font 830 supports
chez SwissLife et 165 chez Abeille, elles changent a chaque publication
(entrees et sorties mensuelles), et une allocation qui cite un support sorti du
contrat est irrecevable chez l assureur.

Usage :
    python3 scripts/extraire-univers-uc.py \\
        --swisslife "~/Downloads/SL-Retraite-et-Epargne-2026-06.xlsx" \\
        --abeille    "~/Downloads/LISTE UC ABEILLE.pdf"

Dependances : openpyxl pour le classeur SwissLife, pdftotext (poppler) pour le
PDF Abeille. Les deux fichiers sources ne sont PAS dans le depot : ils viennent
des assureurs et portent leur propre date de publication, reprise dans le JSON.
"""
import argparse, json, os, re, subprocess, sys, unicodedata
from datetime import date

ISIN = re.compile(r'\b([A-Z]{2}[A-Z0-9]{9}[0-9])\b')


def sl_universe(chemin):
    import openpyxl
    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)

    # 1. Perfs et frais, indexes par ISIN.
    perfs = {}
    if 'Données perf, rétro, frais cour' in wb.sheetnames:
        for row in wb['Données perf, rétro, frais cour'].iter_rows(min_row=4, values_only=True):
            code = str(row[0] or '').strip()
            if not ISIN.fullmatch(code):
                continue
            perfs[code] = {
                'perfNetteAn': nombre(row[7]),      # performance nette N-1, apres frais UC
                'perfNette5AnsAnnualisee': nombre(row[8]),
                'perfFinale': nombre(row[11]),      # apres frais de gestion du contrat
            }

    # 2. La liste elle meme. Les libelles de categorie sont des lignes a une
    #    seule cellule : « Monétaire EUR », « Actions Japon »…
    supports, categorie = [], None
    for row in wb['Liste SL Strat Premium IA'].iter_rows(min_row=5, values_only=True):
        vals = [c for c in row if c is not None]
        if not vals:
            continue
        code = str(row[0] or '').strip()
        if not ISIN.fullmatch(code):
            if len(vals) <= 2 and code:
                categorie = code.strip()
            continue
        supports.append(nettoyer({
            'isin': code,
            'nom': texte(row[3]),
            'categorie': categorie,
            'label': texte(row[1]),
            'sfdr': texte(row[2]),
            'societeGestion': texte(row[5]),
            'devise': texte(row[6]),
            'sri': entier(row[9]),
            'fraisGestionMax': nombre(row[10]),
            **perfs.get(code, {}),
        }))

    # 3. Sorties du mois : un support sorti reste souscrit chez des clients mais
    #    ne peut plus etre propose. On les garde a part pour pouvoir le dire.
    sorties = []
    if 'Liste des sorties' in wb.sheetnames:
        for row in wb['Liste des sorties'].iter_rows(min_row=4, values_only=True):
            code = str(row[0] or '').strip()
            if not ISIN.fullmatch(code):
                continue
            sorties.append(nettoyer({
                'isin': code, 'nom': texte(row[3]),
                'societeGestion': texte(row[5]), 'motif': texte(row[11]),
            }))
    return supports, sorties


# Titres de page du PDF Abeille : « Liste des supports en unités de compte¹ :
# les Obligations ». C est la seule cle de classement fiable, la mise en page a
# deux colonnes melangeant les sous rubriques a l extraction.
TITRE_ABEILLE = re.compile(r'unit[ée]s? de compte.{0,3}\s*:\s*les\s+([A-Za-zÀ-ÿ ]+)', re.I)


# Le PDF liste deux fois les memes supports : par classe d actifs (la
# cartographie) puis par societe de gestion. Seule la premiere serie porte la
# classe, on la depouille donc en premier pour que chaque support la recoive,
# et la seconde ne sert qu a ramasser ceux qui n apparaissaient nulle part
# ailleurs. « Actions par zones geographiques » et « Actions par secteurs » se
# rangent sous « Actions », le detail vit dans le nom du support.
def normaliser_classe(c):
    if not c:
        return None
    c = re.sub(r'\s+par\b.*$', '', c).strip()
    return {'Monétaires': 'Monétaire', 'Actions': 'Actions', 'Mixtes': 'Mixtes',
            'Obligations': 'Obligations', 'Spéculatifs': 'Spéculatifs'}.get(c, c)


def abeille_universe(chemin):
    txt = subprocess.run(['pdftotext', '-layout', chemin, '-'],
                         capture_output=True, text=True, check=True).stdout
    supports, vus = [], set()
    pages = txt.split('\f')
    classees = [(p, TITRE_ABEILLE.search(p)) for p in pages]
    # La classe se releve d abord sur la seule presence de l ISIN dans une page
    # de la cartographie, sans attendre que la ligne soit lisible : un nom de
    # support qui passe a la ligne cassait le parseur et le support se
    # retrouvait sans classe, retrouve plus loin sur la page de sa societe de
    # gestion, qui n en porte aucune.
    classe_par_isin = {}
    for page, m in classees:
        if not m:
            continue
        c = normaliser_classe(m.group(1).strip().rstrip('¹').strip())
        for mi in ISIN.finditer(page):
            classe_par_isin.setdefault(mi.group(1), c)
    ordre = [(p, m) for p, m in classees if m] + [(p, m) for p, m in classees if not m]
    for page, m in ordre:
        classe = normaliser_classe(m.group(1).strip().rstrip('¹').strip()) if m else None
        for ligne in page.split('\n'):
            m2 = ISIN.search(ligne)
            if not m2:
                continue
            code = m2.group(1)
            reste = ligne[m2.end():].strip()
            # « … <nom> <SRI> <SFDR> <label> » : le SRI est un chiffre isole de
            # 1 a 7, le premier a droite du nom.
            sri = None
            msri = re.search(r'\s([1-7])\s+(Art\.|Article)', reste)
            if msri:
                sri = int(msri.group(1))
                nom = reste[:msri.start()].strip()
            else:
                nom = re.sub(r'\s+(Art\.|Article)\s*\d.*$', '', reste).strip()
            nom = re.sub(r'\s{2,}', ' ', nom).strip()
            if not nom or code in vus:
                continue
            vus.add(code)
            supports.append(nettoyer({
                'isin': code, 'nom': nom, 'categorie': classe_par_isin.get(code) or classe, 'sri': sri,
                'sfdr': 'Art. 9' if 'Art. 9' in reste else 'Art. 8' if 'Art. 8' in reste else 'Art. 6' if 'Art. 6' in reste else None,
            }))
    return supports


def texte(v):
    if v is None:
        return None
    s = unicodedata.normalize('NFC', str(v)).replace('_x000D_', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s or None


def nombre(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).replace('%', '').replace(',', '.').strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return None   # « Création 2025 », « n.d. » : pas de perf, on ne devine pas


def entier(v):
    n = nombre(v)
    return int(n) if n is not None else None


def nettoyer(d):
    return {k: v for k, v in d.items() if v is not None}


def ecrire(cible, contenu):
    os.makedirs(os.path.dirname(cible), exist_ok=True)
    with open(cible, 'w', encoding='utf-8') as f:
        json.dump(contenu, f, ensure_ascii=False, indent=0)
    print(f'{cible} : {len(contenu["supports"])} supports')


if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--swisslife')
    p.add_argument('--abeille')
    p.add_argument('--sortie', default='public/data')
    p.add_argument('--date-swisslife', default=None, help='date de publication de la liste')
    p.add_argument('--date-abeille', default=None)
    a = p.parse_args()

    if a.swisslife:
        supports, sorties = sl_universe(os.path.expanduser(a.swisslife))
        ecrire(f'{a.sortie}/univers-uc-swisslife.json', {
            'partenaire': 'swisslife',
            'sourceFichier': os.path.basename(a.swisslife),
            'publie': a.date_swisslife,
            'extraitLe': date.today().isoformat(),
            'supports': supports,
            'sorties': sorties,
        })
    if a.abeille:
        supports = abeille_universe(os.path.expanduser(a.abeille))
        ecrire(f'{a.sortie}/univers-uc-abeille.json', {
            'partenaire': 'abeille',
            'sourceFichier': os.path.basename(a.abeille),
            'publie': a.date_abeille,
            'extraitLe': date.today().isoformat(),
            'supports': supports,
            'sorties': [],
        })
    if not a.swisslife and not a.abeille:
        p.error('donne au moins --swisslife ou --abeille')
